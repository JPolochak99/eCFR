import re
from io import BytesIO

import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

st.title("CFR to Excel Export")

title = st.text_input("Title", "49")
subtitle = st.text_input("Subtitle", "B")
chapter = st.text_input("Chapter", "I")
subchapter = st.text_input("Subchapter", "C")
part = st.text_input("Part", "173")
subpart = st.text_input("Subpart", "I")
section = st.text_input("Section", "173.435")
date = st.date_input("eCFR date")

if "table_catalog" not in st.session_state:
    st.session_state.table_catalog = []
if "endpoint_found" not in st.session_state:
    st.session_state.endpoint_found = False
if "endpoint_error" not in st.session_state:
    st.session_state.endpoint_error = ""


def fetch_page(title, subtitle, chapter, subchapter, part, subpart, section, date):
    base_url = "https://www.ecfr.gov"
    date_str = date.isoformat()

    url = f"{base_url}/api/versioner/v1/full/{date_str}/title-{title}.xml"
    params = {
        "subtitle": subtitle,
        "chapter": chapter,
        "subchapter": subchapter,
        "part": part,
        "subpart": subpart,
        "section": section,
    }
    headers = {
        "Accept": "application/xml,text/xml,*/*",
        "User-Agent": "Mozilla/5.0",
    }

    response = requests.get(url, headers=headers, params=params, timeout=60)
    response.raise_for_status()
    response.encoding = "utf-8"
    return response.text


def build_table_catalog(xml_text):
    soup = BeautifulSoup(xml_text, "xml")
    tables = soup.find_all("TABLE")

    catalog = []
    for i, table in enumerate(tables, start=1):
        caption = table.find("CAPTION")
        caption_text = caption.get_text(" ", strip=True) if caption else ""

        prev = table.find_previous(
            ["H1", "H2", "H3", "H4", "H5", "H6", "P", "DIV", "B", "STRONG"]
        )
        prev_text = prev.get_text(" ", strip=True) if prev else ""

        title_text = caption_text or prev_text or ""
        title_text = truncate(title_text, max_len=100)


        catalog.append(
            {
                "index": i,
                "title": title_text,
                "html": str(table),
            }
        )

    return catalog


def parse_radionuclide(value):
    if pd.isna(value):
        return value, "", ""

    s = str(value).strip()

    notes = re.findall(r"\(([^)]*)\)", s)
    note_text = "; ".join(n.strip() for n in notes if n.strip())

    core = re.sub(r"\s*\([^)]*\)", "", s).strip()

    m = re.match(r"^([A-Za-z]+)-(\d+)([A-Za-z]*)$", core)
    if not m:
        return s, note_text, core

    symbol = m.group(1)
    mass = m.group(2).zfill(3)
    suffix = m.group(3)

    formatted = f"{mass}{symbol}{suffix}"
    return formatted, note_text, core


def convert_scientific(value):
    if pd.isna(value):
        return value

    s = str(value).strip()

    if s == "" or s.lower() == "unlimited" or "see §" in s:
        return s

    s = s.replace("−", "-").replace("×", "x").replace("X", "x")
    s = re.sub(r"\s+", " ", s)

    m = re.match(r"^([0-9.]+)\s*x\s*10\s*([+-]?\d+)$", s)
    if m:
        mantissa = float(m.group(1))
        exponent = int(m.group(2))
        return mantissa * (10 ** exponent)

    try:
        return float(s)
    except Exception:
        return value

def looks_like_scientific(value):
    if pd.isna(value):
        return False
    s = str(value).strip()
    return bool(re.match(r"^[0-9.]+\s*[×x]\s*10\s*[−-]?\d+$", s) or re.match(r"^[0-9.]+$", s))


def convert_scientific(value):
    if pd.isna(value):
        return value

    s = str(value).strip()

    if s == "" or s.lower() == "unlimited" or "see §" in s:
        return s

    s = s.replace("−", "-").replace("×", "x").replace("X", "x")
    s = re.sub(r"\s+", " ", s)

    m = re.match(r"^([0-9.]+)\s*x\s*10\s*([+-]?\d+)$", s)
    if m:
        mantissa = float(m.group(1))
        exponent = int(m.group(2))
        return mantissa * (10 ** exponent)

    try:
        return float(s)
    except Exception:
        return value

def table_html_to_df(table_html):
    soup = BeautifulSoup(table_html, "xml")
    table = soup.find("TABLE") or soup.find("table")

    if table is None:
        raise ValueError("Could not parse selected table.")

    # Collect all table rows
    body = table.find("TBODY")
    tr_source = body.find_all("TR") if body else table.find_all("TR")

    rows = []
    for tr in tr_source:
        cells = [cell.get_text(" ", strip=True) for cell in tr.find_all(["TD", "TH"])]
        if cells:
            rows.append(cells)

    if not rows:
        raise ValueError("No rows found in the selected table.")

    # Try to find a real header row
    headers = None

    # 1) Prefer THEAD if present
    thead = table.find("THEAD")
    if thead:
        header_tr = thead.find("TR")
        if header_tr:
            headers = [cell.get_text(" ", strip=True) for cell in header_tr.find_all(["TH", "TD"])]

    # 2) Otherwise, use the first TR that contains TH cells
    if headers is None:
        for tr in tr_source:
            th_cells = tr.find_all("TH")
            if th_cells:
                headers = [cell.get_text(" ", strip=True) for cell in tr.find_all(["TH", "TD"])]
                break

    # Clean header values if we found them
    if headers:
        headers = [h if h else f"Column {i}" for i, h in enumerate(headers, start=1)]

        max_cols = max(len(headers), max(len(r) for r in rows))
        headers = headers + [f"Column {i}" for i in range(len(headers) + 1, max_cols + 1)]
        rows = [r + [""] * (max_cols - len(r)) for r in rows]

        # If the header row is also present in the extracted rows, remove it
        first_row = [str(x).strip() for x in rows[0]]
        header_row = [str(x).strip() for x in headers[:len(first_row)]]

        if first_row == header_row:
            rows = rows[1:]

        return pd.DataFrame(rows, columns=headers[:max_cols])

    # Fallback: no reliable header row found
    max_cols = max(len(r) for r in rows)
    rows = [r + [""] * (max_cols - len(r)) for r in rows]
    columns = [f"Column {i}" for i in range(1, max_cols + 1)]

    return pd.DataFrame(rows, columns=columns)

def format_df(df):
    numeric_cols = []

    # Radionuclide column formatting only if it exists
    for col in df.columns:
        if "symbol" in col.lower() and "radionuclide" in col.lower():
            parsed = df[col].apply(parse_radionuclide).apply(pd.Series)
            parsed.columns = [col, "Symbol note", "Symbol core"]

            df = df.drop(columns=[col]).join(parsed)

            first_cols = [col, "Symbol note", "Symbol core"]
            other_cols = [c for c in df.columns if c not in first_cols]
            df = df[first_cols + other_cols]
            break

    # Numeric conversion based on cell content, not hardcoded column names
    for col in df.columns:
        if df[col].dtype == object:
            sample = df[col].dropna().astype(str).head(20)

            if not sample.empty:
                matches = sample.apply(lambda x: bool(re.match(r"^[0-9.]+\s*[×x]?\s*10?\s*[−-]?\d*$", x.strip()))).mean()
                if matches > 0.5:
                    df[col] = df[col].apply(convert_scientific)
                    numeric_cols.append(col)

    return df, numeric_cols


def df_to_excel_bytes(df, sheet_name="CFR", styled=True, numeric_cols=None):
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]

        if styled:
            last_row = ws.max_row
            last_col = ws.max_column
            table_ref = f"A1:{get_column_letter(last_col)}{last_row}"

            excel_table = Table(displayName="CFRTable", ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            excel_table.tableStyleInfo = style
            ws.add_table(excel_table)

            if numeric_cols:
                sci_format = "0.0E+00"
                for col_name in numeric_cols:
                    if col_name in df.columns:
                        col_idx = df.columns.get_loc(col_name) + 1
                        for row in range(2, ws.max_row + 1):
                            ws.cell(row=row, column=col_idx).number_format = sci_format

    buffer.seek(0)
    return buffer.getvalue()


def truncate(text, max_len=50):
    return text if len(text) <= max_len else text[:max_len] + "..."

if st.button("Find Tables"):
    try:
        xml_text = fetch_page(
            title, subtitle, chapter, subchapter, part, subpart, section, date
        )
        catalog = build_table_catalog(xml_text)

        st.session_state.table_catalog = catalog
        st.session_state.endpoint_found = True
        st.session_state.endpoint_error = ""

    except Exception as e:
        st.session_state.table_catalog = []
        st.session_state.endpoint_found = False
        st.session_state.endpoint_error = str(e)
        st.error(f"Failed to find tables: {e}")


if st.session_state.endpoint_error:
    st.info(st.session_state.endpoint_error)

if st.session_state.endpoint_found and st.session_state.table_catalog:
    labels = [
        item["title"] if item["title"] else f"Table {item['index']}"
        for item in st.session_state.table_catalog
    ]

    if len(labels) == 1:
        st.info("1 table found.")
        selected_index = 0
    else:
        st.warning(f"Multiple tables found: {len(labels)}.")
        selected_label = st.selectbox("Available tables", labels)
        selected_index = labels.index(selected_label)

    selected_item = st.session_state.table_catalog[selected_index]

    if selected_item["title"]:
        st.caption(selected_item["title"])

    col1, col2 = st.columns(2)

    with col1:
        if st.button("Download Selected Table (Formatted)"):
            try:
                df = table_html_to_df(selected_item["html"])
                df, numeric_cols = format_df(df)
                excel_bytes = df_to_excel_bytes(
                    df,
                    sheet_name="CFR",
                    styled=True,
                    numeric_cols=numeric_cols,
                )

                st.download_button(
                    "Download Formatted Excel",
                    excel_bytes,
                    file_name=f"section_{section.replace('.', '_')}_formatted.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                st.error(f"Formatted export failed: {e}")

    with col2:
        if st.button("Download Selected Table (Raw)"):
            try:
                df = table_html_to_df(selected_item["html"])
                excel_bytes = df_to_excel_bytes(
                    df,
                    sheet_name="CFR",
                    styled=False,
                    numeric_cols=None,
                )

                st.download_button(
                    "Download Raw Excel",
                    excel_bytes,
                    file_name=f"section_{section.replace('.', '_')}_raw.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                st.error(f"Raw export failed: {e}")