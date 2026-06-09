from __future__ import annotations

import re
from io import BytesIO

import pandas as pd

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

INVALID_SHEET_CHARS = r"[:\\/?*\[\]]"


def sanitize_sheet_name(name: str) -> str:
    cleaned = re.sub(INVALID_SHEET_CHARS, "_", str(name).strip())
    cleaned = cleaned[:31].strip()
    return cleaned or "Sheet1"


def unique_sheet_name(name: str, used_names: set[str]) -> str:
    base = sanitize_sheet_name(name)

    if base not in used_names:
        used_names.add(base)
        return base

    i = 2
    while True:
        suffix = f"_{i}"
        candidate = base[: 31 - len(suffix)] + suffix
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        i += 1


def make_source_key(
    title: str,
    subtitle: str,
    chapter: str,
    subchapter: str,
    part: str,
    subpart: str,
    section: str,
    effective_date_str: str,
    table_index: int,
    mode: str,
) -> str:
    return "|".join(
        [
            title,
            subtitle,
            chapter,
            subchapter,
            part,
            subpart,
            section,
            effective_date_str,
            str(table_index),
            mode,
        ]
    )


def workbook_has_table(workbook_tables, source_key: str) -> bool:
    return any(item.get("source_key") == source_key for item in workbook_tables)


def build_workbook_bytes(workbook_tables) -> bytes:
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        used_names: set[str] = set()

        for idx, item in enumerate(workbook_tables, start=1):
            sheet_name = unique_sheet_name(item["sheet_name"], used_names)
            df = item["df"]

            df.to_excel(writer, index=False, sheet_name=sheet_name)

            ws = writer.book[sheet_name]
            last_row = ws.max_row
            last_col = ws.max_column

            table_ref = f"A1:{get_column_letter(last_col)}{last_row}"

            excel_table = Table(
                displayName=f"Table_{idx}",
                ref=table_ref,
            )
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            excel_table.tableStyleInfo = style
            ws.add_table(excel_table)

    buffer.seek(0)
    return buffer.getvalue()