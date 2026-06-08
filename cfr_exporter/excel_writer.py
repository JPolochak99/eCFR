from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "CFR", styled: bool = True, numeric_cols=None) -> bytes:
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]

        if styled:
            last_row = ws.max_row
            last_col = ws.max_column
            table_ref = f"A1:{get_column_letter(last_col)}{last_row}"

            excel_table = Table(displayName="CFRTable", ref=table_ref)
            excel_table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(excel_table)

            if numeric_cols:
                sci_format = "0.0E+00"
                for col_name in numeric_cols:
                    if col_name in df.columns:
                        col_idx = df.columns.get_loc(col_name) + 1
                        for row in range(2, ws.max_row + 1):
                            ws.cell(row=row, column=col_idx).number_format = sci_format

    buffer.seek(0)
    return buffer.getvalue()
