from __future__ import annotations

import re
from io import BytesIO

import pandas as pd

from .uniqueSheetName import unique_sheet_name
from .autoSizeCols import autosize_worksheet_cols
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


from datetime import datetime


from cfr_exporter.derived_sheet_builder import apply_column_formats
from cfr_exporter.convertToScientific import convert_scientific_columns_for_export

def make_source_key(
    title: str,
    subtitle: str,
    chapter: str,
    subchapter: str,
    part: str,
    subpart: str,
    section: str,
    effective_date_str: str,
    table_index: int,
    mode: str,
) -> str:
    return "|".join(
        [
            title,
            subtitle,
            chapter,
            subchapter,
            part,
            subpart,
            section,
            effective_date_str,
            str(table_index),
            mode,
        ]
    )


def workbook_has_table(workbook_tables, source_key: str) -> bool:
    return any(item.get("source_key") == source_key for item in workbook_tables)


def build_workbook_bytes(workbook_tables, metadata: dict | None = None) -> bytes:
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        used_names: set[str] = set()

        # Metadata sheet first
        meta_df = pd.DataFrame(
            [
                {"Field": "Generated", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
                {"Field": "Workbook Name", "Value": (metadata or {}).get("workbook_name", "")},
                {"Field": "CFR Title", "Value": (metadata or {}).get("title", "")},
                {"Field": "Subtitle", "Value": (metadata or {}).get("subtitle", "")},
                {"Field": "Chapter", "Value": (metadata or {}).get("chapter", "")},
                {"Field": "Subchapter", "Value": (metadata or {}).get("subchapter", "")},
                {"Field": "Part", "Value": (metadata or {}).get("part", "")},
                {"Field": "Subpart", "Value": (metadata or {}).get("subpart", "")},
                {"Field": "Section", "Value": (metadata or {}).get("section", "")},
                {"Field": "eCFR Date", "Value": (metadata or {}).get("effective_date_str", "")},
                {"Field": "Sheets Added", "Value": len(workbook_tables)},
            ]
        )

        meta_sheet_name = unique_sheet_name("Workbook Info", used_names)
        meta_df.to_excel(writer, index=False, sheet_name=meta_sheet_name)

        ws = writer.book[meta_sheet_name]

        last_row = ws.max_row
        last_col = ws.max_column
        table_ref = f"A1:{get_column_letter(last_col)}{last_row}"

        meta_table = Table(displayName="WorkbookInfoTable", ref=table_ref)
        meta_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(meta_table)
        autosize_worksheet_cols(ws)

        # User sheets

        for idx, item in enumerate(workbook_tables, start=1):
            sheet_name = unique_sheet_name(item["sheet_name"], used_names)

            formatting = item.get("formatting", {})

            df = item["df"].copy()
            df = make_unique_columns(df)
            df = convert_scientific_columns_for_export(df, formatting)

            df.to_excel(writer, index=False, sheet_name=sheet_name)

            ws = writer.book[sheet_name]

            apply_column_formats(ws, df, formatting)

            last_row = ws.max_row
            last_col = ws.max_column

            if last_row >= 2 and last_col >= 1:
                table_ref = f"A1:{get_column_letter(last_col)}{last_row}"

                excel_table = Table(
                    displayName=f"Table_{idx}",
                    ref=table_ref,
                )

                excel_table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )

                ws.add_table(excel_table)

            autosize_worksheet_cols(ws)

    buffer.seek(0)
    return buffer.getvalue()


def make_unique_columns(df):
    counts = {}
    new_columns = []

    for col in df.columns:
        col = str(col).strip() or "Column"
        if col not in counts:
            counts[col] = 0
            new_columns.append(col)
        else:
            counts[col] += 1
            new_columns.append(f"{col}_{counts[col] + 1}")

    df = df.copy()
    df.columns = new_columns
    return df